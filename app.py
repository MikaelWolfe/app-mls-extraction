import streamlit as st
import pdfplumber
from openpyxl import load_workbook
import re
from rapidfuzz import fuzz
import tempfile

# === Paramètres Excel ===
EXCEL_MODELE = "Comparables.xlsx"

st.title("📄 Importez vos fiches MLS (immeubles à revenus ou maisons unifamiliales) et les rôles d’évaluation associés.")

# --- Fonctions utilitaires ---
def m2_to_p2(valeur_str):
    if not valeur_str:
        return None
    try:
        valeur_str = valeur_str.replace(" ", "").replace(",", ".")
        valeur_str = re.sub(r"[^\d.]", "", valeur_str)
        if valeur_str == "":
            return None
        m2 = float(valeur_str)
        return round(m2 * (3.2808 ** 2), 2)
    except:
        return None

def normalize_address(addr):
    addr = addr.lower()
    addr = re.sub(r"[.,\-]", " ", addr)
    addr = re.sub(r"\s+", " ", addr).strip()
    return addr

def adresse_match(addr1, addr2, seuil=80):
    num1 = re.match(r"^\d+", addr1)
    num2 = re.match(r"^\d+", addr2)
    if num1 and num2 and num1.group() != num2.group():
        return False
    score = fuzz.token_set_ratio(normalize_address(addr1), normalize_address(addr2))
    return score >= seuil

def extract_civic_numbers(addr):
    nums = []
    match = re.match(r"^(\d+)(?:-(\d+))?", addr.strip())
    if match:
        start = int(match.group(1))
        end = int(match.group(2)) if match.group(2) else start
        nums = list(range(start, end+1))
    return nums

def civic_in_mls(addr_other, addr_mls):
    civics_mls = extract_civic_numbers(addr_mls)
    civics_other = extract_civic_numbers(addr_other)
    return any(c in civics_mls for c in civics_other)

# === Charger le workbook ===
wb = load_workbook(EXCEL_MODELE, keep_vba=True)
ws = wb["Feuil1"]

col1, col2 = st.columns(2)

# -------------------
# Colonne gauche → PDFs MLS
# -------------------
with col1:
    st.header("💰 Fiches MLS")
    uploaded_pdfs_mls = st.file_uploader(
        "Glissez vos fichiers PDF MLS", 
        type=["pdf"], 
        accept_multiple_files=True,
        key="upload_mls"
    )

    mls_adresses = []
    mls_rows = []
    mls_nums = []

    start_row = 11
    total_fiches = 0

    if uploaded_pdfs_mls:
        with st.expander("Afficher plus"):
            for uploaded_pdf in uploaded_pdfs_mls:
                sections = []
                section_temp = []
                first_page_found = False

                with pdfplumber.open(uploaded_pdf) as pdf:
                    for page in pdf.pages:
                        texte = page.extract_text()
                        if not texte:
                            continue

                        if not first_page_found and re.search(r"Page\s+1\s+de\s+\d+", texte):
                            first_page_found = True

                        if first_page_found:
                            section_temp.append(texte)
                            if re.search(r"Page\s+(\d+)\s+de\s+\1", texte):
                                sections.append("\n".join(section_temp))
                                section_temp = []

                    if section_temp:
                        sections.append("\n".join(section_temp))

                for section in sections:
                    lignes = section.splitlines()

                    # --- Extraction adresse ---
                    texte_final = "Non trouvé"
                    for i, ligne in enumerate(lignes):
                        if "$" in ligne and i + 1 < len(lignes):
                            texte_final = lignes[i + 1].strip()
                            break

                    # --- Extraction prix ---
                    prix = "Non trouvé"
                    for ligne in lignes:
                        match = re.search(r"\d[\d\s,.]*\s*\$", ligne)
                        if match:
                            prix = match.group(0).strip()
                            break

                    # --- Extraction MLS ---
                    mls_num = "Non trouvé"
                    for ligne in lignes:
                        match = re.search(r"No Centris\s*[:\s]*([\d]+)\s*\(", ligne, re.IGNORECASE)
                        if match:
                            mls_num = match.group(1).strip()
                            break

                    if mls_num in mls_nums:
                        st.warning(f"⚠️ Doublon MLS {mls_num} trouvé → fiche ignorée")
                        continue

                    mls_adresses.append(texte_final)
                    mls_rows.append(start_row)
                    mls_nums.append(mls_num)

                    ws[f"O{start_row}"] = texte_final
                    ws[f"P{start_row}"] = prix

                    st.write(f"📌 Fiche MLS {total_fiches+1} → Rangée Excel {start_row} → {texte_final} → Prix: {prix} → MLS: {mls_num}")

                    start_row += 1
                    total_fiches += 1

# -------------------
# Colonne droite → Autres PDFs
# -------------------
with col2:
    st.header("📄 Évaluations foncières")
    uploaded_pdfs_other = st.file_uploader(
        "Glissez vos autres fichiers PDF", 
        type=["pdf"], 
        accept_multiple_files=True,
        key="upload_other"
    )

    if uploaded_pdfs_other:
        st.write(f"{len(uploaded_pdfs_other)} fichiers téléchargés pour traitement spécial")

        with st.expander("Afficher plus"):
            for idx, uploaded_pdf in enumerate(uploaded_pdfs_other, start=1):
                adresse_other = "Non trouvé"
                superficie = None
                evaluation = None
                aire_etages = None

                with pdfplumber.open(uploaded_pdf) as pdf:
                    for page in pdf.pages:
                        lignes = page.extract_text().splitlines()
                        for ligne in lignes:
                            if "Adresse" in ligne:
                                match = re.search(r"Adresse\s*:\s*(.*)", ligne)
                                if match:
                                    adresse_other = match.group(1).strip()
                            if "Superficie" in ligne:
                                match = re.search(r"Superficie\s*:\s*([\d\s.,$]+)", ligne)
                                if match:
                                    superficie = m2_to_p2(match.group(1).strip())
                            if "Valeur de l'immeuble" in ligne:
                                match = re.search(r"Valeur de l'immeuble\s*:\s*([\d\s.,$]+)", ligne)
                                if match:
                                    evaluation = match.group(1).strip()
                            if "Aire d'étages" in ligne:
                                match = re.search(r"Aire d'étages\s*:\s*([\d\s.,$]+)", ligne)
                                if match:
                                    aire_etages = m2_to_p2(match.group(1).strip())

                correspondance = "Non"
                rangée_mls = None
                mls_number_matched = None
                for i, mls_addr in enumerate(mls_adresses):
                    if adresse_match(adresse_other, mls_addr, seuil=80) or civic_in_mls(adresse_other, mls_addr):
                        correspondance = "Oui"
                        rangée_mls = mls_rows[i]
                        mls_number_matched = mls_nums[i]
                        break

                st.subheader(f"📌 Fichier {idx}")
                st.success(f"Adresse trouvée : {adresse_other}")
                if evaluation:
                    st.write(f"💰 Évaluation municipale trouvée : {evaluation}")
                st.write(f"Ressemblance avec MLS ? {correspondance}")
                if correspondance == "Oui":
                    st.write(f"Comparaison: PDF autre → '{adresse_other}' VS MLS → '{mls_adresses[i]}' (Rangée Excel: {rangée_mls}) → MLS: {mls_number_matched}")

                if rangée_mls:
                    prev_superficie = ws[f"R{rangée_mls}"].value or 0
                    prev_evaluation = ws[f"T{rangée_mls}"].value or 0
                    prev_aire = ws[f"Q{rangée_mls}"].value or 0

                    if superficie:
                        ws[f"R{rangée_mls}"] = prev_superficie + superficie
                    if evaluation:
                        eval_val = float(re.sub(r"[^\d.]", "", evaluation))
                        ws[f"T{rangée_mls}"] = prev_evaluation + eval_val
                    if aire_etages:
                        ws[f"Q{rangée_mls}"] = prev_aire + aire_etages

# -------------------
# Sauvegarde finale
# -------------------
with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as tmp:
    wb.save(tmp.name)
    tmp.seek(0)
    st.download_button(
        "⬇️ Télécharger Excel complet (MLS + Autres PDFs)",
        data=open(tmp.name, "rb").read(),
        file_name="Resultat_Complet_MLS_Autres.xlsm",
        mime="application/vnd.ms-excel.sheet.macroEnabled.12"
    )

